from fastapi import FastAPI, APIRouter, Request, HTTPException, Depends, Query, UploadFile, File
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.exceptions import RequestValidationError
from fastapi.exception_handlers import request_validation_exception_handler
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import unicodedata
from pathlib import Path
from typing import List, Optional
from datetime import datetime, timezone, date
from io import BytesIO
import openpyxl
from pydantic import BaseModel

from models import (
    UserInDB, UserCreate, UserUpdate, UserOut,
    TeacherInDB, TeacherCreate, TeacherUpdate, TeacherOut,
    ClassTypeInDB, ClassTypeCreate, ClassTypeUpdate, ClassTypeOut,
    ClassroomInDB, ClassroomCreate, ClassroomUpdate, ClassroomOut,
    CourseInDB, CourseCreate, CourseUpdate, CourseOut, Schedule,
    ClassSessionInDB, ClassSessionCreate, ClassSessionUpdate, ClassSessionOut,
    StudentInDB, StudentCreate, StudentUpdate, StudentOut,
    AttendanceInDB, AttendanceBulkItem, AttendanceOut, StudentWithAttendance,
    BillingInDB, BillingCreate, BillingUpdate, BillingOut,
    LoginRequest, LoginResponse,
)
from auth import (
    hash_password, verify_password, create_token, decode_token,
    get_current_user, require_role, set_auth_cookie, clear_auth_cookie,
    COOKIE_NAME,
)
from serializers import serialize_doc, serialize_list

def normalize(s):
    return unicodedata.normalize('NFD', str(s).strip()).encode('ascii', 'ignore').decode('utf-8').lower()

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ.get('DB_NAME', 'gestion_clases')]

app = FastAPI(title="Sistema de Gestión de Clases", version="2.0.0")
api = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError):
    logger.warning(
        "Request validation error on %s %s: %s",
        request.method,
        request.url.path,
        exc.errors(),
    )
    return await request_validation_exception_handler(request, exc)


# ─── Startup: Seed SUPERUSER ───
@app.on_event("startup")
async def startup_event():
    existing = await db.users.find_one({"email": "admin@sistema.com"})
    if not existing:
        seed_user = UserInDB(
            name="Super Administrador",
            email="admin@sistema.com",
            password_hash=hash_password("Admin1234!"),
            role="SUPERUSER",
        )
        doc = seed_user.model_dump()
        doc["created_at"] = doc["created_at"].isoformat()
        await db.users.insert_one(doc)
        logger.info("Seed SUPERUSER creado: admin@sistema.com")
    # Create indexes
    await db.users.create_index("email", unique=True)
    await db.users.create_index("id", unique=True)
    await db.teachers.create_index("id", unique=True)
    await db.class_types.create_index("id", unique=True)
    await db.classrooms.create_index("id", unique=True)
    await db.students.create_index("id", unique=True)
    await db.students.create_index("course_id")
    await db.billing.create_index("id", unique=True)
    # Courses
    await db.courses.create_index("id", unique=True)
    # Class sessions
    await db.classes.create_index("id", unique=True)
    await db.classes.create_index("course_id")
    await db.classes.create_index("date")
    # Attendance
    await db.attendance.create_index("id", unique=True)
    await db.attendance.create_index([("class_id", 1), ("student_id", 1)], unique=True)
    logger.info("Base de datos inicializada correctamente")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()


# ─── AUTH ───
@api.post("/auth/login")
async def login(req: LoginRequest):
    user = await db.users.find_one({"email": req.email, "active": True}, {"_id": 0})
    if not user or not verify_password(req.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    token = create_token(user["id"], user["role"], user["email"])
    user_out = UserOut(**serialize_doc(user))
    response = JSONResponse(content={"message": "Login exitoso", "user": user_out.model_dump(mode="json")})
    set_auth_cookie(response, token)
    return response


@api.post("/auth/logout")
async def logout():
    response = JSONResponse(content={"message": "Sesión cerrada"})
    clear_auth_cookie(response)
    return response


@api.get("/auth/me")
async def get_me(request: Request):
    user_data = get_current_user(request)
    user = await db.users.find_one({"id": user_data["userId"], "active": True}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return UserOut(**serialize_doc(user)).model_dump(mode="json")


# ─── USERS (SUPERUSER only) ───
@api.get("/users")
async def list_users(request: Request, user=Depends(require_role("SUPERUSER"))):
    users = await db.users.find({"active": True}, {"_id": 0}).to_list(1000)
    return [UserOut(**serialize_doc(u)).model_dump(mode="json") for u in users]


@api.post("/users")
async def create_user(data: UserCreate, user=Depends(require_role("SUPERUSER"))):
    existing = await db.users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="El email ya está registrado")
    if data.role not in ["ADMIN", "SUPERUSER"]:
        raise HTTPException(status_code=400, detail="Rol inválido")
    new_user = UserInDB(
        name=data.name, email=data.email,
        password_hash=hash_password(data.password), role=data.role,
    )
    doc = new_user.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.users.insert_one(doc)
    return UserOut(**serialize_doc(doc)).model_dump(mode="json")


@api.put("/users/{user_id}")
async def update_user(user_id: str, data: UserUpdate, user=Depends(require_role("SUPERUSER"))):
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if "password" in updates:
        updates["password_hash"] = hash_password(updates.pop("password"))
    if "role" in updates and updates["role"] not in ["ADMIN", "SUPERUSER"]:
        raise HTTPException(status_code=400, detail="Rol inválido")
    if not updates:
        raise HTTPException(status_code=400, detail="No hay datos para actualizar")
    result = await db.users.update_one({"id": user_id, "active": True}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    updated = await db.users.find_one({"id": user_id}, {"_id": 0})
    return UserOut(**serialize_doc(updated)).model_dump(mode="json")


@api.delete("/users/{user_id}")
async def delete_user(user_id: str, user=Depends(require_role("SUPERUSER"))):
    if user["userId"] == user_id:
        raise HTTPException(status_code=400, detail="No puede eliminarse a sí mismo")
    result = await db.users.update_one({"id": user_id, "active": True}, {"$set": {"active": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return {"message": "Usuario desactivado"}


# ─── TEACHERS ───
@api.get("/teachers")
async def list_teachers(request: Request, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    teachers = await db.teachers.find({"active": True}, {"_id": 0}).to_list(1000)
    return serialize_list(teachers)


@api.post("/teachers")
async def create_teacher(data: TeacherCreate, user=Depends(require_role("SUPERUSER"))):
    teacher = TeacherInDB(**data.model_dump())
    doc = teacher.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.teachers.insert_one(doc)
    return serialize_doc(doc)


@api.put("/teachers/{teacher_id}")
async def update_teacher(teacher_id: str, data: TeacherUpdate, user=Depends(require_role("SUPERUSER"))):
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No hay datos para actualizar")
    result = await db.teachers.update_one({"id": teacher_id, "active": True}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Profesor no encontrado")
    updated = await db.teachers.find_one({"id": teacher_id}, {"_id": 0})
    return serialize_doc(updated)


@api.delete("/teachers/{teacher_id}")
async def delete_teacher(teacher_id: str, user=Depends(require_role("SUPERUSER"))):
    result = await db.teachers.update_one({"id": teacher_id, "active": True}, {"$set": {"active": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Profesor no encontrado")
    return {"message": "Profesor desactivado"}


# ─── CLASS TYPES ───
@api.get("/class-types")
async def list_class_types(request: Request, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    types = await db.class_types.find({"active": True}, {"_id": 0}).to_list(1000)
    return serialize_list(types)


@api.post("/class-types")
async def create_class_type(data: ClassTypeCreate, user=Depends(require_role("SUPERUSER"))):
    ct = ClassTypeInDB(**data.model_dump())
    doc = ct.model_dump()
    await db.class_types.insert_one(doc)
    return serialize_doc(doc)


@api.put("/class-types/{type_id}")
async def update_class_type(type_id: str, data: ClassTypeUpdate, user=Depends(require_role("SUPERUSER"))):
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No hay datos para actualizar")
    result = await db.class_types.update_one({"id": type_id, "active": True}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tipo de clase no encontrado")
    updated = await db.class_types.find_one({"id": type_id}, {"_id": 0})
    return serialize_doc(updated)


@api.delete("/class-types/{type_id}")
async def delete_class_type(type_id: str, user=Depends(require_role("SUPERUSER"))):
    result = await db.class_types.update_one({"id": type_id, "active": True}, {"$set": {"active": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Tipo de clase no encontrado")
    return {"message": "Tipo de clase desactivado"}


# ─── CLASSROOMS ───
@api.get("/classrooms")
async def list_classrooms(request: Request, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    rooms = await db.classrooms.find({"active": True}, {"_id": 0}).to_list(1000)
    return serialize_list(rooms)


@api.get("/classrooms/{room_id}")
async def get_classroom(room_id: str, request: Request, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    room = await db.classrooms.find_one({"id": room_id, "active": True}, {"_id": 0})
    if not room:
        raise HTTPException(status_code=404, detail="Salón no encontrado")
    return serialize_doc(room)


@api.post("/classrooms")
async def create_classroom(data: ClassroomCreate, user=Depends(require_role("SUPERUSER"))):
    room = ClassroomInDB(**data.model_dump())
    doc = room.model_dump()
    await db.classrooms.insert_one(doc)
    return serialize_doc(doc)


@api.put("/classrooms/{room_id}")
async def update_classroom(room_id: str, data: ClassroomUpdate, user=Depends(require_role("SUPERUSER"))):
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No hay datos para actualizar")
    result = await db.classrooms.update_one({"id": room_id, "active": True}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Salón no encontrado")
    updated = await db.classrooms.find_one({"id": room_id}, {"_id": 0})
    return serialize_doc(updated)


@api.delete("/classrooms/{room_id}")
async def delete_classroom(room_id: str, user=Depends(require_role("SUPERUSER"))):
    result = await db.classrooms.update_one({"id": room_id, "active": True}, {"$set": {"active": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Salón no encontrado")
    return {"message": "Salón desactivado"}


@api.get("/classrooms/{room_id}/schedules")
async def get_room_schedules(room_id: str, date: Optional[str] = None, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    room = await db.classrooms.find_one({"id": room_id, "active": True}, {"_id": 0})
    if not room:
        raise HTTPException(status_code=404, detail="Salón no encontrado")
    query = {"classroom_id": room_id, "active": True}
    if date:
        query["date"] = date
    sessions = await db.classes.find(query, {"_id": 0}).to_list(1000)
    return [await enrich_session(s) for s in sessions]


# ─── COURSES ───
async def enrich_course(course_doc):
    """Augment course doc with student count. Teacher/class_type info is no longer stored on the course."""
    doc = serialize_doc(course_doc)
    # teacher_id and class_type_id no longer exist; enrich with None for UI compatibility
    doc["teacher_name"] = None
    doc["class_type_name"] = None
    count = await db.students.count_documents({"course_id": doc["id"], "active": True})
    doc["student_count"] = count
    return doc


@api.get("/courses")
async def list_courses(
    request: Request,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=200),
    user=Depends(require_role("ADMIN", "SUPERUSER")),
):
    courses = await db.courses.find({"active": True}, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    return [await enrich_course(c) for c in courses]


@api.get("/courses/{course_id}")
async def get_course(course_id: str, request: Request, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    course = await db.courses.find_one({"id": course_id, "active": True}, {"_id": 0})
    if not course:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    return await enrich_course(course)


@api.get("/courses/{course_id}/students")
async def get_course_students(course_id: str, request: Request, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    course = await db.courses.find_one({"id": course_id, "active": True}, {"_id": 0})
    if not course:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    students = await db.students.find({"course_id": course_id, "active": True}, {"_id": 0}).to_list(1000)
    return serialize_list(students)


@api.get("/courses/{course_id}/classes")
async def get_course_classes(course_id: str, request: Request, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    course = await db.courses.find_one({"id": course_id, "active": True}, {"_id": 0})
    if not course:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    sessions = await db.classes.find({"course_id": course_id, "active": True}, {"_id": 0}).to_list(1000)
    return [await enrich_session(s) for s in sessions]


@api.post("/courses")
async def create_course(data: CourseCreate, user=Depends(require_role("SUPERUSER"))):
    # Create course with new schema (no teacher_id/class_type_id)
    new_course = CourseInDB(**data.model_dump())
    doc = new_course.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.courses.insert_one(doc)
    return await enrich_course(doc)


@api.put("/courses/{course_id}")
async def update_course(course_id: str, data: CourseUpdate, user=Depends(require_role("SUPERUSER"))):
    updates = {}
    for k, v in data.model_dump().items():
        if v is not None:
            if isinstance(v, Schedule):
                updates[k] = v.model_dump()
            else:
                updates[k] = v
    if not updates:
        raise HTTPException(status_code=400, detail="No hay datos para actualizar")
    result = await db.courses.update_one({"id": course_id, "active": True}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    updated = await db.courses.find_one({"id": course_id}, {"_id": 0})
    return await enrich_course(updated)


@api.delete("/courses/{course_id}")
async def delete_course(course_id: str, user=Depends(require_role("SUPERUSER"))):
    result = await db.courses.update_one({"id": course_id, "active": True}, {"$set": {"active": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    return {"message": "Curso desactivado"}


# ─── CLASS SESSIONS ───
def times_overlap(s1_start, s1_end, s2_start, s2_end):
    """Check if two time ranges overlap."""
    return s1_start < s2_end and s2_start < s1_end


async def enrich_session(session_doc):
    """Add course_name, classroom_name, teacher_name, class_type_name to session doc."""
    doc = serialize_doc(session_doc)
    course = await db.courses.find_one({"id": doc.get("course_id")}, {"_id": 0})
    doc["course_name"] = course["name"] if course else None
    room = await db.classrooms.find_one({"id": doc.get("classroom_id")}, {"_id": 0})
    doc["classroom_name"] = room["name"] if room else None
    teacher = await db.teachers.find_one({"id": doc.get("teacher_id")}, {"_id": 0})
    doc["teacher_name"] = teacher["name"] if teacher else None
    class_type = await db.class_types.find_one({"id": doc.get("class_type_id")}, {"_id": 0})
    doc["class_type_name"] = class_type["name"] if class_type else None
    return doc


@api.get("/classes")
async def list_sessions(
    request: Request,
    course_id: Optional[str] = None,
    date: Optional[str] = None,
    recovered: Optional[bool] = None,
    user=Depends(require_role("ADMIN", "SUPERUSER")),
):
    query = {"active": True}
    if course_id:
        query["course_id"] = course_id
    if date:
        query["date"] = date
    if recovered is not None:
        query["recovered"] = recovered
    sessions = await db.classes.find(query, {"_id": 0}).to_list(1000)
    return [await enrich_session(s) for s in sessions]


@api.get("/classes/upcoming")
async def upcoming_sessions(request: Request, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    sessions = await db.classes.find({"date": {"$gte": today}, "active": True}, {"_id": 0}).to_list(1000)
    enriched = [await enrich_session(s) for s in sessions]
    enriched.sort(key=lambda x: (x.get("date", ""), x.get("start_time", "")))
    return enriched


@api.get("/classes/{session_id}")
async def get_session(session_id: str, request: Request, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    session = await db.classes.find_one({"id": session_id, "active": True}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    return await enrich_session(session)


@api.post("/classes")
async def create_session(data: ClassSessionCreate, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    # Validate course exists
    course = await db.courses.find_one({"id": data.course_id, "active": True})
    if not course:
        raise HTTPException(status_code=400, detail="Curso no encontrado")
    # Validate teacher exists
    teacher = await db.teachers.find_one({"id": data.teacher_id, "active": True})
    if not teacher:
        raise HTTPException(status_code=400, detail="Profesor no encontrado")
    # Validate class type exists
    class_type = await db.class_types.find_one({"id": data.class_type_id, "active": True})
    if not class_type:
        raise HTTPException(status_code=400, detail="Tipo de clase no encontrado")
    # Validate classroom exists
    room = await db.classrooms.find_one({"id": data.classroom_id, "active": True})
    if not room:
        raise HTTPException(status_code=400, detail="Salón no encontrado")
    # Warn about classroom conflict if not recovered
    if not data.recovered:
        existing = await db.classes.find(
            {"classroom_id": data.classroom_id, "date": data.date, "active": True}, {"_id": 0}
        ).to_list(1000)
        for ex in existing:
            if times_overlap(data.start_time, data.end_time, ex["start_time"], ex["end_time"]):
                raise HTTPException(
                    status_code=409,
                    detail="Conflicto de horario: el salón ya tiene una sesión programada en ese rango",
                )
        # Check teacher conflict
        existing_teacher = await db.classes.find(
            {"teacher_id": data.teacher_id, "date": data.date, "active": True}, {"_id": 0}
        ).to_list(1000)
        for ex in existing_teacher:
            if times_overlap(data.start_time, data.end_time, ex["start_time"], ex["end_time"]):
                raise HTTPException(
                    status_code=409,
                    detail="Conflicto de horario: el profesor ya tiene una sesión asignada en ese horario",
                )
    new_session = ClassSessionInDB(**data.model_dump())
    doc = new_session.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.classes.insert_one(doc)
    return await enrich_session(doc)


@api.put("/classes/{session_id}")
async def update_session(session_id: str, data: ClassSessionUpdate, user=Depends(require_role("SUPERUSER"))):
    session = await db.classes.find_one({"id": session_id, "active": True}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No hay datos para actualizar")
    
    new_date = updates.get("date", session.get("date"))
    new_start = updates.get("start_time", session.get("start_time"))
    new_end = updates.get("end_time", session.get("end_time"))
    new_room = updates.get("classroom_id", session.get("classroom_id"))
    new_teacher = updates.get("teacher_id", session.get("teacher_id"))
    is_recovered = updates.get("recovered", session.get("recovered", False))
    
    if not is_recovered:
        existing = await db.classes.find(
            {"classroom_id": new_room, "date": new_date, "active": True, "id": {"$ne": session_id}}, {"_id": 0}
        ).to_list(1000)
        for ex in existing:
            if times_overlap(new_start, new_end, ex["start_time"], ex["end_time"]):
                raise HTTPException(
                    status_code=409,
                    detail="Conflicto de horario: el salón ya tiene una sesión programada en ese rango",
                )
        existing_teacher = await db.classes.find(
            {"teacher_id": new_teacher, "date": new_date, "active": True, "id": {"$ne": session_id}}, {"_id": 0}
        ).to_list(1000)
        for ex in existing_teacher:
            if times_overlap(new_start, new_end, ex["start_time"], ex["end_time"]):
                raise HTTPException(
                    status_code=409,
                    detail="Conflicto de horario: el profesor ya tiene una sesión asignada en ese horario",
                )
    
    result = await db.classes.update_one({"id": session_id, "active": True}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    updated = await db.classes.find_one({"id": session_id}, {"_id": 0})
    return await enrich_session(updated)


@api.delete("/classes/{session_id}")
async def delete_session(session_id: str, user=Depends(require_role("SUPERUSER"))):
    result = await db.classes.update_one({"id": session_id, "active": True}, {"$set": {"active": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    return {"message": "Sesión desactivada"}


# ─── STUDENTS ───
@api.get("/students")
async def list_students(
    request: Request,
    search: Optional[str] = None,
    course_id: Optional[str] = None,
    active_only: bool = True,
    user=Depends(require_role("ADMIN", "SUPERUSER")),
):
    query = {"active": active_only}
    if course_id:
        query["course_id"] = course_id
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
        ]
    students = await db.students.find(query, {"_id": 0}).to_list(1000)
    result = []
    for s in students:
        doc = serialize_doc(s)
        if doc.get("course_id"):
            course = await db.courses.find_one({"id": doc["course_id"]}, {"_id": 0})
            doc["course_name"] = course["name"] if course else None
        else:
            doc["course_name"] = None
        result.append(doc)
    return result


@api.get("/students/{student_id}")
async def get_student(student_id: str, request: Request, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    student = await db.students.find_one({"id": student_id, "active": True}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")
    doc = serialize_doc(student)
    if doc.get("course_id"):
        course = await db.courses.find_one({"id": doc["course_id"]}, {"_id": 0})
        doc["course_name"] = course["name"] if course else None
    else:
        doc["course_name"] = None
    return doc


def is_minor(birth_date: date) -> bool:
    today = date.today()
    age = today.year - birth_date.year - (
        (today.month, today.day) < (birth_date.month, birth_date.day)
    )
    return age < 18


@api.post("/students")
async def create_student(data: StudentCreate, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    # Validate course if provided
    if data.course_id:
        course = await db.courses.find_one({"id": data.course_id, "active": True})
        if not course:
            raise HTTPException(status_code=400, detail="Curso no encontrado")
        # Check max_students
        count = await db.students.count_documents({"course_id": data.course_id, "active": True})
        if count >= course["max_students"]:
            raise HTTPException(status_code=400, detail="El curso ha alcanzado el límite máximo de alumnos")
    
    # Validate guardian based on age
    guardian = data.guardian
    if data.birth_date:
        birth = datetime.strptime(data.birth_date, "%Y-%m-%d").date()
        if not is_minor(birth):
            guardian = None
    
    student_data = data.model_dump()
    student_data["guardian"] = guardian.model_dump() if guardian else None
    student = StudentInDB(**student_data)
    doc = student.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.students.insert_one(doc)
    return serialize_doc(doc)


@api.put("/students/{student_id}")
async def update_student(student_id: str, data: StudentUpdate, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    # Get existing student for comparison
    existing = await db.students.find_one({"id": student_id, "active": True})
    if not existing:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")
    
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No hay datos para actualizar")
    
    # Validate and process guardian based on age
    if "guardian" in updates or "birth_date" in updates:
        new_birth_date = updates.get("birth_date", existing.get("birth_date"))
        guardian_in_payload = "guardian" in updates
        new_guardian = updates.get("guardian", existing.get("guardian"))
        
        if new_birth_date:
            birth = datetime.strptime(new_birth_date, "%Y-%m-%d").date()
            if not is_minor(birth):
                updates["guardian"] = None
            elif guardian_in_payload:
                updates["guardian"] = new_guardian.model_dump() if new_guardian else None
        else:
            if guardian_in_payload and new_guardian:
                updates["guardian"] = new_guardian.model_dump() if new_guardian else None
    
    # Validate new course if changing
    if "course_id" in updates and updates["course_id"]:
        course = await db.courses.find_one({"id": updates["course_id"], "active": True})
        if not course:
            raise HTTPException(status_code=400, detail="Curso no encontrado")
        count = await db.students.count_documents({"course_id": updates["course_id"], "active": True})
        if count >= course["max_students"]:
            raise HTTPException(status_code=400, detail="El curso ha alcanzado el límite máximo de alumnos")
    result = await db.students.update_one({"id": student_id, "active": True}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")
    updated = await db.students.find_one({"id": student_id}, {"_id": 0})
    return serialize_doc(updated)


@api.delete("/students/{student_id}")
async def delete_student(student_id: str, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    result = await db.students.update_one({"id": student_id, "active": True}, {"$set": {"active": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")
    return {"message": "Alumno desactivado"}


@api.post("/students/import/excel")
async def import_students_excel(
    request: Request,
    file: Optional[UploadFile] = File(None),
    user=Depends(require_role("ADMIN", "SUPERUSER")),
):
    if file is None:
        content_type = request.headers.get("content-type")
        raise HTTPException(
            status_code=400,
            detail=(
                "No se recibió el archivo. Debe enviarse como multipart/form-data con el campo 'file'. "
                f"Content-Type recibido: {content_type}"
            ),
        )

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx o .xls")
    
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        sheet = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al leer el archivo: {str(e)}")
    
    headers = [normalize(h) for h in next(sheet.iter_rows(min_row=1, max_row=1), [])]
    col_index = {normalize(h): i for i, h in enumerate(headers)}
    
    required_cols = ['nombre']
    for col in required_cols:
        if col not in col_index:
            raise HTTPException(status_code=400, detail=f"Columna '{col}' faltante en el Excel")
    
    created = 0
    errors = []
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        try:
            name = row[col_index['nombre']].value
            if not name:
                errors.append({"row": row_idx, "reason": "nombre es obligatorio"})
                continue
            
            name = str(name).strip()
            birth_date = None
            if 'fecha de nacimiento' in col_index:
                bd = row[col_index['fecha de nacimiento']].value
                if bd:
                    if isinstance(bd, datetime):
                        birth_date = bd.strftime('%Y-%m-%d')
                    else:
                        birth_date = str(bd).strip()
            
            email = None
            if 'email' in col_index:
                e = row[col_index['email']].value
                if e:
                    email = str(e).strip()
            
            phone = None
            if 'telefono' in col_index:
                p = row[col_index['telefono']].value
                if p:
                    phone = str(p).strip()
            
            course_id = None
            if 'curso' in col_index or 'mes de inicio' in col_index:
                course_name = None
                if 'curso' in col_index:
                    cn = row[col_index['curso']].value
                    if cn:
                        course_name = str(cn).strip()
                if not course_name and 'mes de inicio' in col_index:
                    sm = row[col_index['mes de inicio']].value
                    if sm:
                        course_name = str(sm).strip()
                
                if course_name:
                    course = await db.courses.find_one({
                        "$or": [
                            {"name": course_name},
                            {"start_month": course_name}
                        ],
                        "active": True
                    })
                    if course:
                        count = await db.students.count_documents({"course_id": course["id"], "active": True})
                        if count < course["max_students"]:
                            course_id = course["id"]
                        else:
                            errors.append({"row": row_idx, "reason": f"curso '{course_name}' sin cupos disponibles"})
                            continue
                    else:
                        errors.append({"row": row_idx, "reason": f"curso '{course_name}' no encontrado o inactivo"})
                        continue
            
            guardian = None
            # Solo crear guardian si hay fecha de nacimiento y el alumno es menor
            if birth_date and ('nombre tutor' in col_index or 'telefono tutor' in col_index):
                try:
                    birth = datetime.strptime(birth_date, "%Y-%m-%d").date()
                    if is_minor(birth):
                        g_name = row[col_index['nombre tutor']].value if 'nombre tutor' in col_index else None
                        g_phone = row[col_index['telefono tutor']].value if 'telefono tutor' in col_index else None
                        
                        g_name_str = str(g_name).strip() if g_name else ""
                        g_phone_str = str(g_phone).strip() if g_phone else ""
                        
                        # Solo crear guardian si AMBOS campos tienen datos
                        if g_name_str and g_phone_str:
                            guardian = {
                                "name": g_name_str,
                                "phone": g_phone_str,
                            }
                except:
                    pass  # Si falla el parsing de fecha, no incluir guardian
            
            student_data = {
                "name": name,
                "email": email,
                "phone": phone,
                "birth_date": birth_date,
                "guardian": guardian,
                "course_id": course_id,
            }
            
            student = StudentInDB(**student_data)
            doc = student.model_dump()
            doc["created_at"] = doc["created_at"].isoformat()
            await db.students.insert_one(doc)
            created += 1
            
        except Exception as e:
            errors.append({"row": row_idx, "reason": str(e)})
    
    return {"created": created, "failed": len(errors), "errors": errors}


@api.get("/students/{student_id}/attendance")
async def get_student_attendance(student_id: str, request: Request, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    records = await db.attendance.find({"student_id": student_id}, {"_id": 0}).to_list(1000)
    result = []
    for r in records:
        doc = serialize_doc(r)
        session = await db.classes.find_one({"id": r["class_id"]}, {"_id": 0})
        if session:
            doc["class_date"] = session["date"]
            doc["start_time"] = session["start_time"]
            doc["end_time"] = session["end_time"]
            course = await db.courses.find_one({"id": session["course_id"]}, {"_id": 0})
            doc["course_name"] = course["name"] if course else None
        result.append(doc)
    return result


@api.get("/students/{student_id}/billing")
async def get_student_billing(student_id: str, request: Request, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    bills = await db.billing.find({"student_id": student_id}, {"_id": 0}).to_list(1000)
    student = await db.students.find_one({"id": student_id}, {"_id": 0})
    result = []
    for b in bills:
        doc = serialize_doc(b)
        doc["student_name"] = student["name"] if student else "Desconocido"
        result.append(doc)
    return result


# ─── ATTENDANCE ───
@api.get("/attendance/class/{class_id}")
async def get_class_attendance(
    class_id: str, request: Request,
    user=Depends(require_role("ADMIN", "SUPERUSER")),
):
    session = await db.classes.find_one({"id": class_id, "active": True}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    # Get active students from the course
    students = await db.students.find(
        {"course_id": session["course_id"], "active": True}, {"_id": 0}
    ).to_list(1000)
    # Get existing attendance for this session
    existing = await db.attendance.find({"class_id": class_id}, {"_id": 0}).to_list(1000)
    att_map = {a["student_id"]: serialize_doc(a) for a in existing}
    result = []
    for s in students:
        att = att_map.get(s["id"])
        result.append({
            "student_id": s["id"],
            "student_name": s["name"],
            "present": att["present"] if att else None,
            "notes": att.get("notes", "") if att else "",
            "attendance_id": att["id"] if att else None,
        })
    return result


@api.post("/attendance/class/{class_id}")
async def record_attendance(
    class_id: str,
    items: List[AttendanceBulkItem],
    user=Depends(require_role("ADMIN", "SUPERUSER")),
):
    session = await db.classes.find_one({"id": class_id, "active": True}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    saved = []
    now = datetime.now(timezone.utc).isoformat()
    for item in items:
        # Validate student is active in the course
        student = await db.students.find_one(
            {"id": item.student_id, "course_id": session["course_id"], "active": True}
        )
        if not student:
            raise HTTPException(
                status_code=400,
                detail=f"Alumno {item.student_id} no pertenece al curso o no está activo",
            )
        # Upsert attendance
        await db.attendance.update_one(
            {"class_id": class_id, "student_id": item.student_id},
            {
                "$set": {
                    "present": item.present,
                    "notes": item.notes,
                    "recorded_at": now,
                },
                "$setOnInsert": {
                    "class_id": class_id,
                    "student_id": item.student_id,
                    "id": f"{class_id}_{item.student_id}",
                },
            },
            upsert=True,
        )
        saved.append(item.student_id)
    return {"message": f"Asistencia registrada para {len(saved)} alumnos", "student_ids": saved}


# ─── BILLING ───
@api.get("/billing")
async def list_billing(
    request: Request,
    status: Optional[str] = None,
    student_id: Optional[str] = None,
    user=Depends(require_role("ADMIN", "SUPERUSER")),
):
    query = {}
    if status:
        query["status"] = status
    if student_id:
        query["student_id"] = student_id
    bills = await db.billing.find(query, {"_id": 0}).to_list(1000)
    result = []
    for b in bills:
        doc = serialize_doc(b)
        student = await db.students.find_one({"id": b["student_id"]}, {"_id": 0})
        doc["student_name"] = student["name"] if student else "Desconocido"
        result.append(doc)
    return result


@api.get("/billing/{bill_id}")
async def get_billing(bill_id: str, request: Request, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    bill = await db.billing.find_one({"id": bill_id}, {"_id": 0})
    if not bill:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    doc = serialize_doc(bill)
    student = await db.students.find_one({"id": bill["student_id"]}, {"_id": 0})
    doc["student_name"] = student["name"] if student else "Desconocido"
    return doc


@api.post("/billing")
async def create_billing(data: BillingCreate, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    student = await db.students.find_one({"id": data.student_id, "active": True})
    if not student:
        raise HTTPException(status_code=404, detail="Alumno no encontrado")
    bill = BillingInDB(**data.model_dump())
    doc = bill.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.billing.insert_one(doc)
    result = serialize_doc(doc)
    result["student_name"] = student["name"]
    return result


@api.put("/billing/{bill_id}")
async def update_billing(bill_id: str, data: BillingUpdate, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="No hay datos para actualizar")
    if "status" in updates and updates["status"] not in ["PENDING", "PAID", "OVERDUE", "CANCELLED"]:
        raise HTTPException(status_code=400, detail="Estado inválido")
    result = await db.billing.update_one({"id": bill_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    updated = await db.billing.find_one({"id": bill_id}, {"_id": 0})
    doc = serialize_doc(updated)
    student = await db.students.find_one({"id": updated["student_id"]}, {"_id": 0})
    doc["student_name"] = student["name"] if student else "Desconocido"
    return doc


@api.delete("/billing/{bill_id}")
async def delete_billing(bill_id: str, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    result = await db.billing.update_one({"id": bill_id}, {"$set": {"status": "CANCELLED"}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Factura no encontrada")
    return {"message": "Factura cancelada"}


@api.post("/billing/mark-overdue")
async def mark_overdue_billing(user=Depends(require_role("ADMIN", "SUPERUSER"))):
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    result = await db.billing.update_many(
        {"status": "PENDING", "due_date": {"$lt": today}},
        {"$set": {"status": "OVERDUE"}},
    )
    return {"message": f"{result.modified_count} facturas marcadas como vencidas"}


# ─── DASHBOARD ───
@api.get("/dashboard")
async def get_dashboard(request: Request, user=Depends(require_role("ADMIN", "SUPERUSER"))):
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    active_courses = await db.courses.count_documents({"active": True})
    active_students = await db.students.count_documents({"active": True})
    active_teachers = await db.teachers.count_documents({"active": True})
    # Today's sessions
    today_sessions = await db.classes.find({"date": today, "active": True}, {"_id": 0}).to_list(100)
    today_classes = []
    for s in today_sessions:
        enriched = await enrich_session(s)
        today_classes.append(enriched)
    today_classes.sort(key=lambda x: x.get("start_time", ""))
    # Pending billing
    pending_bills = await db.billing.count_documents({"status": "PENDING"})
    overdue_bills = await db.billing.count_documents({"status": "OVERDUE"})
    return {
        "active_courses": active_courses,
        "active_students": active_students,
        "active_teachers": active_teachers,
        "today_classes": today_classes,
        "pending_bills": pending_bills,
        "overdue_bills": overdue_bills,
    }


# ─── EXPORT ENDPOINTS ───
from exports import (
    export_students_csv, export_students_pdf,
    export_classes_csv, export_classes_pdf,
    export_billing_csv, export_billing_pdf,
    export_attendance_csv, export_attendance_pdf,
    export_teachers_csv, export_teachers_pdf,
)


@api.get("/export/students")
async def export_students(
    request: Request,
    format: str = Query("csv", regex="^(csv|pdf)$"),
    user=Depends(require_role("ADMIN", "SUPERUSER")),
):
    students = await db.students.find({"active": True}, {"_id": 0}).to_list(10000)
    if format == "pdf":
        buf = export_students_pdf(students)
        return StreamingResponse(buf, media_type="application/pdf",
                                 headers={"Content-Disposition": "attachment; filename=alumnos.pdf"})
    else:
        buf = export_students_csv(students)
        return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": "attachment; filename=alumnos.csv"})


@api.get("/export/classes")
async def export_courses_report(
    request: Request,
    format: str = Query("csv", regex="^(csv|pdf)$"),
    user=Depends(require_role("ADMIN", "SUPERUSER")),
):
    courses = await db.courses.find({"active": True}, {"_id": 0}).to_list(10000)
    enriched = [await enrich_course(c) for c in courses]
    if format == "pdf":
        buf = export_classes_pdf(enriched)
        return StreamingResponse(buf, media_type="application/pdf",
                                 headers={"Content-Disposition": "attachment; filename=cursos.pdf"})
    else:
        buf = export_classes_csv(enriched)
        return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": "attachment; filename=cursos.csv"})


@api.get("/export/billing")
async def export_billing_report(
    request: Request,
    format: str = Query("csv", regex="^(csv|pdf)$"),
    status: Optional[str] = None,
    user=Depends(require_role("ADMIN", "SUPERUSER")),
):
    query = {}
    if status:
        query["status"] = status
    bills = await db.billing.find(query, {"_id": 0}).to_list(10000)
    for b in bills:
        student = await db.students.find_one({"id": b["student_id"]}, {"_id": 0})
        b["student_name"] = student["name"] if student else "Desconocido"
    if format == "pdf":
        buf = export_billing_pdf(bills)
        return StreamingResponse(buf, media_type="application/pdf",
                                 headers={"Content-Disposition": "attachment; filename=facturacion.pdf"})
    else:
        buf = export_billing_csv(bills)
        return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": "attachment; filename=facturacion.csv"})


@api.get("/export/attendance/{class_id}")
async def export_attendance_report(
    class_id: str,
    request: Request,
    format: str = Query("csv", regex="^(csv|pdf)$"),
    user=Depends(require_role("ADMIN", "SUPERUSER")),
):
    session = await db.classes.find_one({"id": class_id, "active": True}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    # Get active students from course + their attendance
    students = await db.students.find(
        {"course_id": session["course_id"], "active": True}, {"_id": 0}
    ).to_list(1000)
    existing = await db.attendance.find({"class_id": class_id}, {"_id": 0}).to_list(1000)
    att_map = {a["student_id"]: a for a in existing}
    records = []
    for s in students:
        att = att_map.get(s["id"])
        records.append({
            "student_name": s["name"],
            "present": att["present"] if att else None,
            "notes": att.get("notes", "") if att else "",
        })
    # Enrich session info
    course = await db.courses.find_one({"id": session["course_id"]}, {"_id": 0})
    room = await db.classrooms.find_one({"id": session["classroom_id"]}, {"_id": 0})
    session_info = {
        "class_name": course["name"] if course else "-",
        "date": session["date"],
        "start_time": session["start_time"],
        "end_time": session["end_time"],
        "classroom_name": room["name"] if room else "-",
    }

    if format == "pdf":
        buf = export_attendance_pdf(records, session_info)
        return StreamingResponse(buf, media_type="application/pdf",
                                 headers={"Content-Disposition": "attachment; filename=asistencia.pdf"})
    else:
        buf = export_attendance_csv(records, session_info)
        return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": "attachment; filename=asistencia.csv"})


@api.get("/export/teachers")
async def export_teachers_report(
    request: Request,
    format: str = Query("csv", regex="^(csv|pdf)$"),
    user=Depends(require_role("ADMIN", "SUPERUSER")),
):
    teachers = await db.teachers.find({"active": True}, {"_id": 0}).to_list(10000)
    if format == "pdf":
        buf = export_teachers_pdf(teachers)
        return StreamingResponse(buf, media_type="application/pdf",
                                 headers={"Content-Disposition": "attachment; filename=profesores.pdf"})
    else:
        buf = export_teachers_csv(teachers)
        return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv",
                                 headers={"Content-Disposition": "attachment; filename=profesores.csv"})


# Include router + CORS
app.include_router(api)

external_api = APIRouter(prefix="/api/external")


class ExternalStudentCreate(BaseModel):
    name: str
    email: Optional[str] = None
    phone: Optional[str] = None
    birth_date: Optional[str] = None
    course_id: Optional[str] = None


@external_api.post("/students")
async def create_external_student(
    data: ExternalStudentCreate,
    request: Request,
):
    api_key = os.environ.get('EXTERNAL_API_KEY')
    received_key = request.headers.get('X-API-Key')
    
    if not api_key or api_key != received_key:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
    course_id = data.course_id
    if course_id:
        course = await db.courses.find_one({"id": course_id, "active": True})
        if not course:
            raise HTTPException(status_code=400, detail="Curso no encontrado o inactivo")
        count = await db.students.count_documents({"course_id": course_id, "active": True})
        if count >= course["max_students"]:
            raise HTTPException(status_code=400, detail="El curso ha alcanzado el límite máximo de alumnos")
    
    student_data = {
        "name": data.name,
        "email": data.email,
        "phone": data.phone,
        "birth_date": data.birth_date,
        "course_id": course_id,
    }
    
    student = StudentInDB(**student_data)
    doc = student.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.students.insert_one(doc)
    
    return {
        "id": doc["id"],
        "name": doc["name"],
        "created_at": doc["created_at"]
    }

app.include_router(external_api, tags=["external"])

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)
